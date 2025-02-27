import streamlit as st
import pandas as pd
import openpyxl
import zipfile
import os
import io
from modules.procesador_entrevistas import procesar_archivos_adjuntos, generar_reporte

# Configuración de directorios
TEMPLATE_PATH = "templates/plantilla_ficha_evaluacion_EIB.xlsx"
OUTPUT_DIR = "output/"

# Crear directorio de salida si no existe
os.makedirs(OUTPUT_DIR, exist_ok=True)

def procesar_archivos(lista_df):
    """Procesa la lista de entrevistados y genera archivos Excel personalizados."""
    archivos_por_entrevistador = {}
    
    for index, row in lista_df.iterrows():
        nro_doc = row['NRO_DOC']
        nombres = f"{row['APATERNO']} {row['AMATERNO']}, {row['PRIMER NOMBRE']} {row['SEGUNDO NOMBRE']}"
        entrevistador = row['ENTREVISTADOR']
        
        # Leer la plantilla
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        ws = wb["Evaluación de entrevista"]
        
        # Modificar las celdas
        ws["B2"] = nro_doc
        ws["B3"] = nombres
        ws["C33"] = entrevistador
        
        # Definir el nombre del archivo
        file_name = f"ENTREV_{entrevistador}-{nro_doc}-{nombres}.xlsx"
        file_path = os.path.join(OUTPUT_DIR, file_name)
        
        # Guardar el archivo
        wb.save(file_path)
        
        # Agrupar por entrevistador
        if entrevistador not in archivos_por_entrevistador:
            archivos_por_entrevistador[entrevistador] = []
        archivos_por_entrevistador[entrevistador].append(file_path)
    
    return archivos_por_entrevistador

def generar_zips(archivos_por_entrevistador):
    """Genera archivos ZIP con las evaluaciones de cada entrevistador."""
    zip_files = []
    
    for entrevistador, archivos in archivos_por_entrevistador.items():
        zip_name = f"{OUTPUT_DIR}{entrevistador}.zip"
        with zipfile.ZipFile(zip_name, 'w') as zipf:
            for file in archivos:
                zipf.write(file, os.path.basename(file))
        zip_files.append(zip_name)
    
    return zip_files

def main():
    st.title("Gestor de Entrevistas y Fichas - EIB")
    
    tab1, tab2 = st.tabs(["📄 Generar ficha de entrevistas", "📊 Procesar fichas"])
    
    with tab1:
        generar_evaluaciones()
    
    with tab2:
        procesar_entrevistas()

def generar_evaluaciones():
    st.header("Generación de Evaluaciones")
    uploaded_file = st.file_uploader("Subir lista de entrevistados (Excel)", type=["xlsx"], key="upload_eval")
    
    if uploaded_file is not None:
        df = pd.read_excel(uploaded_file, dtype=str)
        required_columns = {'NRO_DOC', 'APATERNO', 'AMATERNO', 'PRIMER NOMBRE', 'SEGUNDO NOMBRE', 'ENTREVISTADOR'}
        if not required_columns.issubset(df.columns):
            st.error("El archivo no contiene todas las columnas requeridas.")
            return
        
        N = len(df)
        st.write(f"Cantidad de postulantes: {N}")
        archivos_por_entrevistador = procesar_archivos(df)
        zip_files = generar_zips(archivos_por_entrevistador)
        
        if "zip_files" not in st.session_state:
            st.session_state.zip_files = zip_files

        for zip_file in st.session_state.zip_files:
            with open(zip_file, "rb") as f:
                zip_bytes = f.read()
            st.download_button(
                label=f"Descargar {os.path.basename(zip_file)}",
                data=zip_bytes,
                file_name=os.path.basename(zip_file),
                mime="application/zip",
                key=f"download_{os.path.basename(zip_file)}"
            )

def procesar_entrevistas():
    st.header("Procesador de Entrevistas")
    uploaded_files = st.file_uploader("Subir archivos de entrevistas (Excel o ZIP)", type=["xlsx", "zip"], accept_multiple_files=True, key="upload_proc")
    
    if uploaded_files:
        resultados = procesar_archivos_adjuntos(uploaded_files)
        if resultados:
            df_reporte = generar_reporte(resultados)
            st.dataframe(df_reporte)

            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df_reporte.to_excel(writer, index=False, sheet_name="Reporte")
                buffer.seek(0)

            st.download_button(
                label="Descargar Reporte Consolidado",
                data=buffer.getvalue(),
                file_name="reporte_entrevistas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

if __name__ == "__main__":
    main()
