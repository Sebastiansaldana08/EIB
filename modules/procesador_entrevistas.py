import pandas as pd
import openpyxl
import zipfile
import os
import io

def extraer_datos_excel(file_path):
    """Extrae la información clave de la hoja 'Evaluación de entrevista' de un archivo Excel."""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb["Evaluación de entrevista"]
        
        nro_documento = ws["B2"].value
        nombres = ws["B3"].value
        entrevistador = ws["C33"].value
        nota_total = ws["C34"].value
        
        return {
            "NRO_DOC": nro_documento,
            "NOMBRES": nombres,
            "ENTREVISTADOR": entrevistador,
            "NOTA_TOTAL": nota_total
        }
    except Exception as e:
        return None

def procesar_archivos_adjuntos(uploaded_files):
    """Procesa archivos Excel y ZIP para extraer los datos de cada evaluación."""
    resultados = []
    
    for uploaded_file in uploaded_files:
        if uploaded_file.name.endswith(".xlsx"):
            # Procesar archivo Excel individual
            with io.BytesIO(uploaded_file.read()) as temp_file:
                with open("temp.xlsx", "wb") as f:
                    f.write(temp_file.getbuffer())
                data = extraer_datos_excel("temp.xlsx")
                if data:
                    resultados.append(data)
                os.remove("temp.xlsx")
        
        elif uploaded_file.name.endswith(".zip"):
            # Procesar archivo ZIP
            with zipfile.ZipFile(io.BytesIO(uploaded_file.read()), 'r') as zip_ref:
                for file_name in zip_ref.namelist():
                    if file_name.endswith(".xlsx"):
                        with zip_ref.open(file_name) as extracted_file:
                            with open("temp.xlsx", "wb") as f:
                                f.write(extracted_file.read())
                            data = extraer_datos_excel("temp.xlsx")
                            if data:
                                resultados.append(data)
                            os.remove("temp.xlsx")
    
    return resultados

def generar_reporte(resultados):
    """Genera un DataFrame con el reporte consolidado de evaluaciones."""
    df = pd.DataFrame(resultados)
    return df
